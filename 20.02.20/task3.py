import openpyxl

book = openpyxl.Workbook()

ws1 = book.create_sheet('Sheet 1')  # insert at the end (default)
ws1.title = 'I fixed the title'

ws1['A1'] = 1000  # access by  key of a worksheet
ws1['A2'] = 100

ws1.cell(row=2, column=2).value = 100  # access by row and column

d = ws1.cell(row=3, column=122, value=100)
d.value = 322

ws1.append((1, 2))  # insert at the end of the worksheet
ws1.append([1, 3, 4])  # value should be a list, tuple, range of generator or dict

book.save('sample.xlsx')
